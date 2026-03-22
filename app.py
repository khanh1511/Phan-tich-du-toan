import streamlit as st
import pandas as pd
import io

st.set_page_config(page_title="Tổng hợp Dự toán", layout="wide", page_icon="📊")

st.title("📊 Ứng dụng Xử lý & Tổng hợp File Dự toán")
st.markdown("""
Ứng dụng này giúp bạn:
1. Tra cứu **Phân loại công tác** từ **Thư viện DM** dựa vào **Mã số**.
2. Chèn cột **Phân loại công tác** vào ngay sau cột Mã số của **Dữ liệu DA**.
3. Tổng hợp khối lượng, thành tiền theo từng **Phân loại công tác**.
""")

col1, col2 = st.columns(2)

with col1:
    st.subheader("1. Tải lên Thư viện Định mức")
    lib_file = st.file_uploader("Chọn file Thư viện DM (.xlsx)", type=['xlsx', 'xls'], key="lib")

with col2:
    st.subheader("2. Tải lên Dữ liệu Dự án")
    data_file = st.file_uploader("Chọn file Dữ liệu DA (.xlsx)", type=['xlsx', 'xls'], key="data")

if lib_file and data_file:
    try:
        # Tải dữ liệu
        df_lib = pd.read_excel(lib_file, engine='openpyxl')
        df_data = pd.read_excel(data_file, engine='openpyxl')
        
        st.success("✅ Tải file thành công! Hãy chọn các cột tương ứng bên dưới để xử lý.")
        
        # Cấu hình chọn cột
        st.header("⚙️ Cấu hình Cột Dữ liệu")
        
        col_lib1, col_lib2 = st.columns(2)
        with col_lib1:
            lib_ma_so_options = df_lib.columns.tolist()
            default_lib_ma_so = next((c for c in lib_ma_so_options if "Mã" in str(c) or "mã" in str(c).lower()), lib_ma_so_options[0] if lib_ma_so_options else "")
            lib_ma_so = st.selectbox("Cột Mã số (trong Thư viện)", options=lib_ma_so_options, index=lib_ma_so_options.index(default_lib_ma_so) if default_lib_ma_so in lib_ma_so_options else 0)
            
        with col_lib2:
            lib_phan_loai_options = df_lib.columns.tolist()
            default_lib_phan_loai = next((c for c in lib_phan_loai_options if "Phân loại" in str(c) or "Loại" in str(c)), lib_phan_loai_options[-1] if lib_phan_loai_options else "")
            lib_phan_loai = st.selectbox("Cột Phân loại (trong Thư viện)", options=lib_phan_loai_options, index=lib_phan_loai_options.index(default_lib_phan_loai) if default_lib_phan_loai in lib_phan_loai_options else (len(lib_phan_loai_options)-1 if lib_phan_loai_options else 0))
            
        col_data1, col_data2, col_data3, col_data4 = st.columns(4)
        with col_data1:
            data_ma_so_options = df_data.columns.tolist()
            default_data_ma_so = next((c for c in data_ma_so_options if "Mã" in str(c) or "mã" in str(c).lower()), data_ma_so_options[0] if data_ma_so_options else "")
            data_ma_so = st.selectbox("Cột Mã số (Dữ liệu DA)", options=data_ma_so_options, index=data_ma_so_options.index(default_data_ma_so) if default_data_ma_so in data_ma_so_options else 0)
            
        with col_data2:
            data_don_vi_options = df_data.columns.tolist()
            default_data_dv = next((c for c in data_don_vi_options if "Đơn vị" in str(c) or "đơn vị" in str(c).lower() or "đvt" in str(c).lower()), data_don_vi_options[0] if data_don_vi_options else "")
            data_don_vi = st.selectbox("Cột Đơn vị (Dữ liệu DA)", options=data_don_vi_options, index=data_don_vi_options.index(default_data_dv) if default_data_dv in data_don_vi_options else 0)

        with col_data3:
            data_khoi_luong_options = df_data.columns.tolist()
            default_data_kl = next((c for c in data_khoi_luong_options if "Khối lượng" in str(c) or "KL" in str(c) or "khối lượng" in str(c).lower()), data_khoi_luong_options[-2] if len(data_khoi_luong_options)>1 else (data_khoi_luong_options[0] if data_khoi_luong_options else ""))
            data_khoi_luong = st.selectbox("Cột Khối lượng (Dữ liệu DA)", options=data_khoi_luong_options, index=data_khoi_luong_options.index(default_data_kl) if default_data_kl in data_khoi_luong_options else 0)
            
        with col_data4:
            data_thanh_tien_options = df_data.columns.tolist()
            default_data_tt = next((c for c in data_thanh_tien_options if "Thành tiền" in str(c) or "Thành tiền" in str(c) or "Tiền" in str(c) or "tt" in str(c).lower()), data_thanh_tien_options[-1] if data_thanh_tien_options else "")
            data_thanh_tien = st.selectbox("Cột Thành tiền (Dữ liệu DA)", options=data_thanh_tien_options, index=data_thanh_tien_options.index(default_data_tt) if default_data_tt in data_thanh_tien_options else 0)

        # Nút xử lý
        if st.button("🚀 Thực thi Xử lý Dữ liệu", type="primary", use_container_width=True):
            with st.spinner("Đang tra cứu và tổng hợp dữ liệu..."):
                # Chuẩn hóa cột khóa để join chính xác
                df_data[data_ma_so] = df_data[data_ma_so].astype(str).str.strip()
                df_lib_lookup = df_lib[[lib_ma_so, lib_phan_loai]].dropna(subset=[lib_ma_so]).copy()
                df_lib_lookup[lib_ma_so] = df_lib_lookup[lib_ma_so].astype(str).str.strip()
                
                # Xóa bản ghi trùng lặp trong thư viện (lấy bản ghi đầu tiên)
                df_lib_lookup = df_lib_lookup.drop_duplicates(subset=[lib_ma_so], keep='first')
                
                # Đổi tên cột phân loại để dễ chèn và gộp
                pl_col_name = 'Phân loại công tác'
                df_lib_lookup = df_lib_lookup.rename(columns={lib_phan_loai: pl_col_name})
                
                # Gộp dữ liệu (merge left)
                merged_df = pd.merge(df_data, df_lib_lookup, left_on=data_ma_so, right_on=lib_ma_so, how="left")
                
                # Bỏ cột khóa thư viện nếu bị thừa
                if lib_ma_so != data_ma_so and lib_ma_so in merged_df.columns and lib_ma_so != pl_col_name:
                    merged_df = merged_df.drop(columns=[lib_ma_so])
                
                # Kiểm tra xem tên 'Phân loại công tác' có trùng với tên cột nào sẵn có trong data DA không
                # Chú ý pd.merge có thể thêm hậu tố _x, _y nếu trùng tên.
                if f"{pl_col_name}_y" in merged_df.columns:
                    merged_df = merged_df.rename(columns={f"{pl_col_name}_y": pl_col_name})
                    if f"{pl_col_name}_x" in merged_df.columns:
                         merged_df = merged_df.drop(columns=[f"{pl_col_name}_x"])
                
                # Sắp xếp lại thứ tự cột
                if pl_col_name in merged_df.columns:
                    cols = merged_df.columns.tolist()
                    cols.remove(pl_col_name)
                    ma_so_idx = cols.index(data_ma_so)
                    cols.insert(ma_so_idx + 1, pl_col_name)
                    merged_df = merged_df[cols]
                
                # Điền giá trị cho công tác chưa phân loại
                merged_df[pl_col_name] = merged_df[pl_col_name].fillna('Chưa phân loại')
                
                # Hàm chuyển đổi đơn vị
                def normalize_and_convert(val, unit):
                    val = pd.to_numeric(val, errors='coerce')
                    if pd.isna(val):
                        val = 0.0
                    
                    if pd.isna(unit):
                        return val, "Không rõ"
                        
                    unit_str = str(unit).strip().lower()
                    
                    # Thể tích -> m3
                    if unit_str in ['m3', 'mét khối', 'm khối']: return val, 'm3'
                    elif unit_str in ['lít', 'lit', 'l']: return val / 1000, 'm3'
                    elif unit_str in ['cm3']: return val / 1000000, 'm3'
                    
                    # Diện tích -> m2
                    elif unit_str in ['m2', 'mét vuông', 'm vuông']: return val, 'm2'
                    elif unit_str in ['cm2']: return val / 10000, 'm2'
                    elif unit_str in ['ha', 'hecta']: return val * 10000, 'm2'
                    
                    # Trọng lượng -> kg
                    elif unit_str in ['kg', 'kilogram', 'kí', 'ký']: return val, 'kg'
                    elif unit_str in ['g', 'gram']: return val / 1000, 'kg'
                    elif unit_str in ['tấn']: return val * 1000, 'kg'
                    elif unit_str in ['tạ']: return val * 100, 'kg'
                    elif unit_str in ['yến']: return val * 10, 'kg'
                    
                    # Chiều dài -> m
                    elif unit_str in ['m', 'mét', 'met']: return val, 'm'
                    elif unit_str in ['km', 'kilomet']: return val * 1000, 'm'
                    elif unit_str in ['cm', 'centimet']: return val / 100, 'm'
                    elif unit_str in ['mm', 'millimet']: return val / 1000, 'm'
                    
                    # Giữ nguyên
                    return val, unit

                # Chuyển đổi và quy đổi đơn vị
                converted_kl = []
                converted_dv = []
                for _, row in merged_df.iterrows():
                    kl, dv = normalize_and_convert(row[data_khoi_luong], row.get(data_don_vi, ""))
                    converted_kl.append(kl)
                    converted_dv.append(dv)
                
                col_kl_quydoi = f"{data_khoi_luong} (Quy đổi)"
                col_dv_chuan = "Đơn vị (chuẩn)"
                merged_df[col_kl_quydoi] = converted_kl
                merged_df[col_dv_chuan] = converted_dv
                
                merged_df[data_thanh_tien] = pd.to_numeric(merged_df[data_thanh_tien], errors='coerce').fillna(0)
                
                # Tính tổng hợp
                agg_df = merged_df.groupby([pl_col_name, col_dv_chuan], as_index=False).agg({
                    col_kl_quydoi: 'sum',
                    data_thanh_tien: 'sum'
                })
                
                # Hàm tô đậm các dòng có chứa dấu *
                def highlight_star_row(row):
                    is_star = any('*' in str(val) for val in row if pd.notna(val))
                    return ['font-weight: bold' if is_star else '' for _ in row]
                
                # Hiển thị trên màn hình
                st.markdown("---")
                tab1, tab2 = st.tabs(["📄 Chi tiết Dự án (Đã kèm Phân loại)", "📊 Tổng hợp theo Phân loại"])
                
                with tab1:
                    st.dataframe(merged_df.style.apply(highlight_star_row, axis=1), use_container_width=True)
                
                with tab2:
                    st.dataframe(agg_df.style.apply(highlight_star_row, axis=1), use_container_width=True)
                
                # Tạo file Excel để tải xuống
                from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
                
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    merged_df.to_excel(writer, sheet_name="Chi tiết dự án", index=False)
                    agg_df.to_excel(writer, sheet_name="Tổng hợp phân loại", index=False)
                    
                    # Format Excel cho đẹp mắt
                    workbook = writer.book
                    bold_font = Font(bold=True)
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    thin_border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    
                    for sheet_name in workbook.sheetnames:
                        worksheet = workbook[sheet_name]
                        
                        # Format Header (Dòng 1)
                        for cell in worksheet[1]:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                            cell.border = thin_border
                            
                        # Format các dòng dữ liệu
                        for row in worksheet.iter_rows(min_row=2):
                            # Kiểm tra xem dòng có chứa dấu * không
                            is_bold = any('*' in str(cell.value) for cell in row if cell.value is not None)
                            
                            for cell in row:
                                cell.border = thin_border
                                if is_bold:
                                    cell.font = bold_font
                                
                                # Format số lượng / tiền tệ để có dấu phẩy phân cách hàng nghìn
                                if isinstance(cell.value, (int, float)):
                                    # Nếu là số nguyên thì không hiện phần thập phân
                                    cell.number_format = '#,##0.00' if isinstance(cell.value, float) and not cell.value.is_integer() else '#,##0'
                        
                        # Tự động điều chỉnh độ rộng cột
                        for col in worksheet.columns:
                            max_length = 0
                            column = col[0].column_letter
                            for cell in col:
                                try:
                                    if cell.value and len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            worksheet.column_dimensions[column].width = min(max_length + 2, 50)
                
                st.markdown("---")
                st.success("🎉 Xử lý hoàn tất! Nhấn nút bên dưới để tải kết quả về máy.")
                
                col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
                with col_btn2:
                    st.download_button(
                        label="⬇️ TẢI FILE EXCEL KẾT QUẢ",
                        data=output.getvalue(),
                        file_name="Ket_qua_xu_ly_DA.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

    except Exception as e:
        st.error(f"Đã xảy ra lỗi khi đọc/xử lý file: {e}")
        st.info("Vui lòng kiểm tra lại định dạng file hoặc các cột được chọn.")
